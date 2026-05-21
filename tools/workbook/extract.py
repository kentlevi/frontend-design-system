"""Extract content workbook sheets to per-sheet JSON files.

Reads the team Content Workbook (Excel) and normalizes column headers across
the various sheets. The header canonicalization handles legacy variants
(`EN (Marketing) - v1`, `EN (Marketing) - v2`, `EN (Marketing) - Implemented`,
`EN (Marketing) - Latest`, the plain `EN`, the `Text` column on Pop_Up, etc.).

Each sheet becomes one JSON file under the output dir, named after the sheet
with whitespace and special characters collapsed to underscores.

Usage:
    python extract.py --src "<path-to-Musticker - Content (N).xlsx>" --out "<out-dir>"

If --src is omitted, looks for `MUSTICKER_WORKBOOK` in the environment, then a
default of "Musticker - Content.xlsx" in the script directory.
If --out is omitted, writes to the script directory.

Requires:
    pip install openpyxl
"""
import argparse, json, os, re, sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

SKIP_SHEETS = {
    "List of Products", "Copy of Homepage", "Sheet11", "SEO",
}


def canon_header(h: str | None) -> str | None:
    """Map a workbook header cell to a canonical short key, or None to ignore."""
    if not h:
        return None
    n = re.sub(r"\s+", " ", h.strip().lower()).replace("--", "-").strip()
    if "figma content" in n:
        return "figma"
    if n == "text":
        return "figma"  # Pop_Up sheet uses "Text" instead of "Figma Content"
    if n.startswith("en (marketing)"):
        if "v1" in n: return "en_v1"
        if "v2" in n: return "en_v2"
        if "implemented" in n: return "en_impl"
        if "latest" in n: return "en_latest"
        return "en_marketing"
    if n == "en":
        return "en_marketing"
    if n.startswith("kr(translation)"):
        if "v1" in n: return "kr_v1"
        if "v2" in n: return "kr_v2"
        if "implemented" in n: return "kr_impl"
        if "latest" in n: return "kr_latest"
        return "kr_marketing"
    if n == "no.":
        return "no"
    if n == "scope":
        return "scope"
    return None


def safe_sheet_name(name: str) -> str:
    """Turn 'Product Page | Die-Cut Stickers' into 'Product_Page_Die_Cut_Sticke' etc."""
    s = re.sub(r"[^\w\-]+", "_", name).strip("_")
    return s[:30]  # cap length to match historical convention


def extract(src: Path, out: Path) -> None:
    try:
        import openpyxl
    except ImportError:
        print("Missing dependency: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    if not src.exists():
        print(f"Workbook not found: {src}", file=sys.stderr)
        sys.exit(1)

    out.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(src, data_only=True)

    written = 0
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # Sheets put the canonical header on row 2 (row 1 is usually a title).
        header = [str(c).strip() if c else "" for c in rows[1]]
        canon = [canon_header(h) for h in header]
        if "scope" not in canon and "figma" not in canon:
            continue

        records = []
        current_scope = None
        for r in rows[2:]:
            if not any(c is not None and str(c).strip() for c in r):
                continue
            rec = {}
            for i, c in enumerate(canon):
                if not c:
                    continue
                v = r[i] if i < len(r) else None
                rec[c] = str(v).strip() if v is not None and str(v).strip() else None
            if rec.get("scope"):
                current_scope = rec["scope"]
            rec["_scope"] = current_scope
            content_keys = (
                "figma", "en_v1", "en_v2", "en_impl", "en_latest", "en_marketing",
                "kr_v1", "kr_v2", "kr_impl", "kr_latest", "kr_marketing",
            )
            if not any(rec.get(k) for k in content_keys):
                continue
            records.append(rec)

        if not records:
            continue
        out_path = out / f"{safe_sheet_name(name)}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        written += 1
        print(f"wrote {out_path.name} ({len(records)} rows)")
    print(f"\nTotal: {written} sheets")


def main():
    p = argparse.ArgumentParser(description="Extract content workbook to per-sheet JSON.")
    p.add_argument("--src", help="Path to the .xlsx workbook (or set MUSTICKER_WORKBOOK env var).")
    p.add_argument("--out", help="Output directory (default: script directory).")
    args = p.parse_args()

    here = Path(__file__).parent
    src = Path(args.src) if args.src else Path(os.environ.get("MUSTICKER_WORKBOOK", here / "Musticker - Content.xlsx"))
    out = Path(args.out) if args.out else here

    extract(src, out)


if __name__ == "__main__":
    main()
